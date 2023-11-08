import os
import re
from openpyxl.styles import Font
from openpyxl import load_workbook
from fpdf import FPDF

def add_suffix_to_filename(file_path, suffix):
    directory_path = os.path.dirname(file_path)
    filename, extension = os.path.splitext(os.path.basename(file_path))
    updated_filename = filename + suffix + extension
    updated_file_path = os.path.join(directory_path, updated_filename)
    return updated_file_path


def create_sheet(sheet_num: int, workbook):
    work_Sheet = workbook[f'Sheet{str(sheet_num)}']
    return work_Sheet

def convert_to_pdf(workbook, sheet_range, page_size):
    wb = load_workbook(workbook)
    sheets = wb[sheet_range]
    pdf = FPDF()
    pdf.add_page()

    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt="Sheet to PDF", ln=1, align="C")
    pdf.cell(200, 10, txt="Page size: " + page_size, ln=2, align="L")

    for sheet in sheets:
        for row in sheet.iter_rows():
            for cell in row:
                pdf.cell(40, 10, str(cell.value), border=1)
            pdf.ln()
    
    pdf.output('updated_output.pdf')

def getNumData(start_row: int, column: str, workSheet) -> dict:
    start_cell, end_row, data_num = workSheet[column + str(start_row)], start_row, {}
    itr_cell = start_cell
    while itr_cell.value is not None:
        data_num[str(end_row - start_row + 1)] = itr_cell.value
        end_row += 1
        itr_cell = workSheet[column + str(end_row)]
    return data_num
    # length = (end_row - start_row) + 1


def autoFillSum(start_row: int, end_row: int, column: str, workSheet, num_list: list, data2: dict, totals: list):
    total_sum = 0
    for row in range(start_row, end_row + 1):
        cell = workSheet[column + str(row)]
        value = cell.value
        if value is not None:
            str_list = re.findall(r'\d+', value)
            temp_sum, count = 0, 0
            for digit in str_list:
                if digit != '' and digit in num_list:
                    count += 1
                    temp_sum += data2[digit]
                else:
                    temp_sum = 0
                    break
            total_sum += temp_sum
            next_cell = workSheet[chr(ord(column) + 1) + str(row)]
            next_cell.value = temp_sum
    sum_cell = workSheet[chr(ord(column) + 1) + str(end_row + 1)]
    sum_cell.font, sum_cell.value = Font(bold=True), total_sum
    totals.append(total_sum)


def update_excel(my_workbook):
    # my_workbook = openpyxl.load_workbook(file)
    data2 = getNumData(3, 'B', create_sheet(1, my_workbook))

    num_list, totals = [str(num) for num in range(1, len(data2) + 1)], []

    autoFillSum(3, 32, 'B', create_sheet(2, my_workbook), num_list, data2, totals)
    autoFillSum(3, 32, 'E', create_sheet(2, my_workbook), num_list, data2, totals)
    autoFillSum(3, 40, 'B', create_sheet(3, my_workbook), num_list, data2, totals)
    autoFillSum(3, 40, 'E', create_sheet(3, my_workbook), num_list, data2, totals)

    # Grand total
    create_sheet(3, my_workbook)['F43'].font = Font(bold=True, italic=True, size=14)
    create_sheet(3, my_workbook)['E43'].value, create_sheet(3, my_workbook)['F43'].value = "GRAND TOTAL", sum(totals)

    return my_workbook


if __name__ == "__main__":
    print(add_suffix_to_filename("uchits.xlsx", "_updated"))
